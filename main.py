import asyncio
import os
import re
import sqlite3
from datetime import datetime

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    CallbackQuery,
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
from PIL import Image as PILImage

BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("Укажите переменную окружения BOT_TOKEN")

EXCEL_FILE = "report.xlsx"
IMG_DIR = "bot_images"
DB_FILE = "bot_data.db"

CATEGORIES = [
    "Общестроительные_работы",
    "Отопление_вентиляция_кондиционирование",
    "ЭОМ",
    "АПС",
    "АПТ",
]

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())
os.makedirs(IMG_DIR, exist_ok=True)


class BotStates(StatesGroup):
    waiting_for_bti = State()


def init_db():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_msg_id INTEGER,
                date_str TEXT,
                caption TEXT,
                category TEXT,
                bti_room TEXT,
                file_path TEXT
            )
            """
        )
        conn.commit()
        cursor.execute("PRAGMA table_info(messages)")
        columns = [col[1] for col in cursor.fetchall()]
        for column, ddl in (
            ("telegram_msg_id", "ALTER TABLE messages ADD COLUMN telegram_msg_id INTEGER"),
            ("category", "ALTER TABLE messages ADD COLUMN category TEXT"),
            ("bti_room", "ALTER TABLE messages ADD COLUMN bti_room TEXT"),
        ):
            if column not in columns:
                cursor.execute(ddl)
                conn.commit()


def add_to_db(msg_id, date_str, caption, file_path):
    with sqlite3.connect(DB_FILE) as conn:
        conn.execute(
            "INSERT INTO messages (telegram_msg_id, date_str, caption, category, bti_room, file_path) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (msg_id, date_str, caption, "Разное", "Не указано", file_path),
        )
        conn.commit()


def update_caption_in_db(msg_id, new_caption):
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE messages SET caption = ? WHERE telegram_msg_id = ?",
            (new_caption, msg_id),
        )
        conn.commit()
        return cursor.rowcount


def update_category_in_db(msg_id, category_name):
    with sqlite3.connect(DB_FILE) as conn:
        conn.execute(
            "UPDATE messages SET category = ? WHERE telegram_msg_id = ?",
            (category_name, msg_id),
        )
        conn.commit()


def update_bti_in_db(msg_id, bti_text):
    with sqlite3.connect(DB_FILE) as conn:
        conn.execute(
            "UPDATE messages SET bti_room = ? WHERE telegram_msg_id = ?",
            (bti_text, msg_id),
        )
        conn.commit()


def get_all_from_db():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT date_str, bti_room, caption, category, file_path FROM messages"
        )
        return cursor.fetchall()


def clear_db():
    with sqlite3.connect(DB_FILE) as conn:
        conn.execute("DELETE FROM messages")
        conn.commit()


def sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "_", name or "Разное").strip()
    return (cleaned or "Разное")[:31]


def generate_excel(data_rows):
    wb = Workbook()
    wb.remove(wb.active)
    sheets = {}
    target_size = (250, 250)

    for _date_str, bti_room, text, category, img_path in data_rows:
        sheet_name = sanitize_sheet_name(category)
        if sheet_name not in sheets:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Помещение (БТИ)", "Текст (Подпись)", "Фотография"])
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 25
            sheets[sheet_name] = (ws, 2)

        ws, current_row = sheets[sheet_name]
        if os.path.exists(img_path):
            with PILImage.open(img_path) as img:
                img = img.convert("RGB")
                img.thumbnail(target_size)
                img.save(img_path, format="JPEG")
            ws.cell(row=current_row, column=1, value=bti_room)
            ws.cell(row=current_row, column=2, value=text)
            ws.add_image(OpenPyXLImage(img_path), f"C{current_row}")
            ws.row_dimensions[current_row].height = (target_size[1] * 0.75) + 10
        else:
            ws.cell(row=current_row, column=1, value=bti_room)
            ws.cell(row=current_row, column=2, value=f"{text} (Фото удалено)")
        sheets[sheet_name] = (ws, current_row + 1)

    if not wb.sheetnames:
        wb.create_sheet(title="Пусто")
    wb.save(EXCEL_FILE)


def get_categories_keyboard(msg_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text=cat.replace("_", " "),
                    callback_data=f"set_cat:{msg_id}:{idx}",
                )
            ]
            for idx, cat in enumerate(CATEGORIES)
        ]
    )


@dp.message(Command("start"))
async def cmd_start(message: Message):
    await message.reply(
        "1. Отправьте фото с подписью\n"
        "2. Выберите раздел\n"
        "3. Укажите номер помещения по БТИ\n"
        "4. /generate — Excel-отчёт"
    )


@dp.message(F.photo)
async def handle_photo(message: Message, state: FSMContext):
    if await state.get_state() == BotStates.waiting_for_bti:
        await message.reply("Сначала укажите номер БТИ или /cancel")
        return

    photo = message.photo[-1]
    local_path = os.path.join(IMG_DIR, f"{photo.file_unique_id}.jpg")
    await bot.download(photo, destination=local_path)
    add_to_db(
        message.message_id,
        message.date.replace(tzinfo=None).strftime("%d.%m.%Y %H:%M"),
        message.caption or "Без описания",
        local_path,
    )
    await message.reply(
        "Выберите раздел:",
        reply_markup=get_categories_keyboard(message.message_id),
    )


@dp.callback_query(F.data.startswith("set_cat:"))
async def process_category_click(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer("Некорректная кнопка", show_alert=True)
        return
    try:
        msg_id = int(parts[1])
        cat_name = CATEGORIES[int(parts[2])]
    except (ValueError, IndexError):
        await callback.answer("Раздел не найден", show_alert=True)
        return

    update_category_in_db(msg_id, cat_name)
    await state.update_data(current_msg_id=msg_id)
    await state.set_state(BotStates.waiting_for_bti)
    await callback.message.edit_text(
        f"Раздел «{cat_name.replace('_', ' ')}» выбран.\n\n"
        "Отправьте номер помещения по БТИ или `-`"
    )
    await callback.answer()


@dp.message(Command("cancel"))
async def cmd_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.reply("Отменено.")


@dp.message(BotStates.waiting_for_bti)
async def process_bti_text(message: Message, state: FSMContext):
    if not message.text:
        await message.reply("Отправьте текст или `-`")
        return
    msg_id = (await state.get_data()).get("current_msg_id")
    if msg_id:
        update_bti_in_db(msg_id, message.text.strip())
        await message.reply(f"Помещение {message.text.strip()} сохранено.")
    await state.clear()


@dp.edited_message()
async def handle_any_edited_message(message: Message):
    update_caption_in_db(message.message_id, message.caption or "Без описания")


@dp.message(Command("ping"))
async def check_status(message: Message):
    await message.reply(f"Бот на связи. {datetime.now():%d.%m.%Y %H:%M:%S}")


@dp.message(Command("generate"))
async def send_report(message: Message):
    data_rows = get_all_from_db()
    if not data_rows:
        await message.reply("База пуста!")
        return
    status_msg = await message.reply("Формирую Excel...")
    try:
        generate_excel(data_rows)
        await message.reply_document(
            FSInputFile(EXCEL_FILE),
            caption=f"Готово. Фото: {len(data_rows)}",
        )
        for *_, img_path in data_rows:
            if os.path.exists(img_path):
                os.remove(img_path)
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
        clear_db()
    except Exception as e:
        await message.reply(f"Ошибка: {e}")
    finally:
        try:
            await status_msg.delete()
        except Exception:
            pass


async def main():
    init_db()
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
